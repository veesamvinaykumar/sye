from flask import Flask, render_template, request, redirect, flash, session
import openpyxl
import os
from datetime import datetime
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Needed for flash messages
SECRET_ADMIN_CODE = "sye@2025"

EXCEL_FILE = "registration.xlsx"
ADMIN_FILE = "admins.xlsx"

# ✅ Step 1: Create or recover Excel files
def initialize_excel():
    try:
        if not os.path.exists(EXCEL_FILE):
            raise FileNotFoundError
        openpyxl.load_workbook(EXCEL_FILE)
    except (FileNotFoundError, BadZipFile, InvalidFileException):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Student Name", "Age", "Father Name", "Dance Type", "Plan", "Payment", "Timestamp", "Email"])
        wb.save(EXCEL_FILE)

def initialize_admin_excel():
    if not os.path.exists(ADMIN_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Username", "Password"])
        wb.save(ADMIN_FILE)

initialize_excel()
initialize_admin_excel()

@app.route('/')
def home():
    return render_template('index.html')
@app.route('/register')
def register():
    return render_template('register.html')


@app.route('/submit', methods=['POST'])
def submit():
    try:
        name = request.form['student_name']
        age = request.form['student_age']
        father = request.form['father_name']
        dance_type = request.form['dance_type']
        plan = request.form['plan']
        payment = request.form['payment']
        email = request.form.get('email', '')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        student_id = ws.max_row  # Auto ID
        ws.append([student_id, name, age, father, dance_type, plan, payment, timestamp, email])
        wb.save(EXCEL_FILE)
        wb.close()

        return render_template("thankyou.html", name=name, dance_type=dance_type)

    except Exception as e:
        return f"An error occurred: {e}"

@app.route('/view')
def view():
    if 'admin' not in session:
        return redirect('/admin')

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    return render_template("view.html", data=data)

@app.route('/search', methods=['POST'])
def search():
    if 'admin' not in session:
        return redirect('/admin')

    student_name = request.form.get('student_name', '').strip().lower()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    data = [row for row in ws.iter_rows(min_row=2, values_only=True) if student_name in str(row[1]).lower()]
    found = len(data) > 0
    return render_template("view.html", data=data, found=found, searched=True)

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get("username")
        password = request.form.get("password")

        try:
            wb = openpyxl.load_workbook(ADMIN_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == username and row[1] == password:
                    session['admin'] = username
                    return redirect("/view")
            return render_template("admin.html", error="Invalid credentials.")
        except:
            return render_template("admin.html", error="Admin data not found.")

    return render_template("admin.html")

@app.route('/admin-register', methods=['GET', 'POST'])
def admin_register():
    if request.method == 'POST':
        code = request.form['secret_code']
        if code != SECRET_ADMIN_CODE:
            return render_template("admin_register.html", error="Invalid secret code.")

        username = request.form['username']
        password = request.form['password']

        wb = openpyxl.load_workbook(ADMIN_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                return render_template("admin_register.html", error="Username already exists.")

        ws.append([username, password])
        wb.save(ADMIN_FILE)
        wb.close()
        return render_template("admin_register.html", success="Admin registered successfully!")

    return render_template("admin_register.html")



@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect('/')


if __name__ == '__main__':
    app.run(debug=True)


